from selenium import webdriver
import openpyxl
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

workbook = openpyxl.load_workbook('Model.xlsx')
sheet = workbook['Sheet1']
PATH = "C:\Program Files (x86)\chromedriver.exe"
driver = webdriver.Chrome(PATH)

def enterMarks(res):
   try:
       exist = WebDriverWait(driver,15).until(
           EC.presence_of_element_located((By.XPATH, '//*[@id="assessment-mark-entry"]/tbody/tr'))
       )
       if exist:
           rows = len(driver.find_elements_by_xpath('//*[@id="assessment-mark-entry"]/tbody/tr'))

           for z in range(len(res)):
               shifana = res[z]


               for i in range(1,rows+1):
                   if list(shifana.keys())[0] == driver.find_element_by_xpath(f'//*[@id="assessment-mark-entry"]/tbody/tr[{i}]/td[1]').text:
                       temp = list(shifana.values())
                       for j in range(4,21):
                           value = temp[0][j - 3]
                           try:
                               mark = WebDriverWait(driver, 15).until(
                                   EC.presence_of_element_located((By.XPATH,
                                                                   f'/html/body/div[6]/div[2]/div/div[2]/div/div/div[3]/div/div[2]/table/tbody/tr[{i}]/td[{j}]/input'))
                               )

                               mark.send_keys(Keys.CONTROL + "a")
                               mark.send_keys(Keys.BACK_SPACE)

                               mark.send_keys(str(value))
                           except(err):
                               print('error 4: ', err)
       else : print("not exist")
   except(err):
       print('error: ',err)


def openPage(res):
   driver.get("https://sstm.linways.com/staff")
   try:
       uname = WebDriverWait(driver,15).until(
           EC.presence_of_element_located((By.XPATH,"//*[@id='username']"))
       )
       uname.send_keys(Keys.CONTROL + "a")
       uname.send_keys(Keys.BACK_SPACE)
       uname.send_keys({username})
       passw = driver.find_element_by_xpath('//*[@id="password"]')
       passw.send_keys({password})
       signIn = driver.find_element_by_xpath('/html/body/form/div[1]/div[3]/div[2]/div/button')
       signIn.click()
   except(err) :
      print('shit00', err)

   try:
       s2 = WebDriverWait(driver,15).until(
           EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/div[2]/div/div[2]/div[1]/div[2]/div[1]/a/div"))
       )
       s2.click()
   except(err):
       print('error 2: ',err)

   try:
       qnPaper = WebDriverWait(driver,15).until(
           EC.presence_of_element_located((By.XPATH,"//*[@id='htw_course_outcome']/a/b"))
       )
       qnPaper.click()
   except(err):
       print('error 3: ',err)

   try:
       enterMarksButton = WebDriverWait(driver,15).until(
           EC.presence_of_element_located((By.XPATH, "//*[@id='htw_course_outcome_enter_marks']/a"))
       )
       enterMarksButton.click()
   except(err):
       print('error 4: ',err)

   try:
       b2 = WebDriverWait(driver,15).until(
           EC.presence_of_element_located((By.XPATH, "//*[@id='content']/table/tbody/tr[7]/td[5]/a"))
       )
       b2.click()
   except(err):
       print('error 4: ',err)

   enterMarks(res)

res= []
marks = {i:0 for i in range(1,18)}
marks['total_marks'] = 0
#print(marks)
roll = None

for i in range(2,1262):
  if str(sheet.cell(i,1).value).isdigit() or None==sheet.cell(i,1).value==sheet.cell(i,2).value==sheet.cell(i,3).value:
      #save to reslut array if there is
        if roll:
          res.append({roll:marks})
          marks = {i: 0 for i in range(1, 18)}
      #create new dict
        roll = 'FM-'+str(sheet.cell(i,1).value)
        marks[sheet.cell(i, 2).value] = sheet.cell(i, 3).value

  elif sheet.cell(i,2).value == None:
      marks['total_marks'] = sum(marks.values())


  else :
      marks[sheet.cell(i,2).value] = sheet.cell(i,3).value

#
# for x in res:
#     # print(list(x.keys()))
#     # break
#     if 'FM-1751' == list(x.keys())[0]:
#         print(list(x.values()))
#     else: print("no such user")

openPage(res)
# shifana = res[90]
# jack = (list(shifana.values()))
# for i in range(1,19):
#     print(jack[0][i])


